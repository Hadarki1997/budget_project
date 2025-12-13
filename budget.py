from pprint import pprint
import random
import numpy as np
import os
from openpyxl import load_workbook


def clean_expenses(expenses):
    my_list = []
    for exp in expenses:
        amount = exp.get("amount")
        category = exp.get("category")
        payment = exp.get("payment") 

        if amount is None or amount <= 0:
            continue
        if category is None or category.strip() == "":
            continue 
        if isinstance(amount, int):
            amount = float(amount)  

        my_new_expenses = {
            "category": category,
            "amount": amount,
            "payment": payment,
        }
        my_list.append(my_new_expenses)
    return my_list      

def total_by_category(expenses):
    if not expenses:
        print("you have not expenses. good job")
        return {}

    total = {}
    for exp in expenses:
        category = exp["category"]
        amount = exp["amount"]

        if category not in total:
            total[category] = 0

        total[category] += amount

    return total

def categories_above(expenses, threshold):
    s = set()
    totals = total_by_category(expenses)

    for category, amount in totals.items():
        if amount > threshold:
            s.add(category)
    return s

def expense_stats(expenses):

    cleaned = clean_expenses(expenses)  

    if not cleaned:                    
        print("you have not expenses. good job")
        return None, None

    amount_list = []
    for exp in cleaned:
        amount = exp["amount"]
        amount_list.append(amount)

    arr_amount_list = np.array(amount_list)
    ave = np.mean(arr_amount_list)
    hetzion = np.median(arr_amount_list)
    return ave, hetzion 

def simulate_month(num_expenses):
    my_list = []
    list_categories = ["food","rent","fun","transport","other"]
    ranges = {
    "food": (30, 200),
    "rent": (1500, 4000),
    "fun": (20, 300),
    "transport": (10, 80),
    "other": (5, 150)
}
    payment = ["card","cash","bank"]
    for i in range(num_expenses):
        category = random.choice(list_categories)
        amount = int(random.uniform(*ranges[category]))
        payment = random.choice(payment)
        
        new_expense = {
        "category": category,
        "amount": amount,
        "payment": payment }
        my_list.append(new_expense)

    return my_list


class Budget:

    def __init__(self, limit):
        self.limitation = limit
        self.expenses = []

    def load_from_excel(self):
        while True:
            filename = input("Enter Excel file name (including .xlsx): ")
            if os.path.exists(filename):
                break
            print("File not found. Please try again.\n")

        wb = load_workbook(filename)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or (isinstance(row[0], str) and "סך" in row[0]) or row[5] is None:
                continue
            self.add_expense(row[2], row[5], row[14])

    def add_expense(self, category, amount, payment):
        expense = {
            "category": category,
            "amount": amount,
            "payment": payment
        }
        cleaned = clean_expenses([expense])
        if cleaned:
            self.expenses.append(cleaned[0])

    def total_spent(self):
        total = 0
        for exp in self.expenses:
            total += exp["amount"]
        return total

    def remaining(self):
        return self.limitation - self.total_spent()

    def category_report(self):
        categories = total_by_category(self.expenses)
        for category, total in categories.items():
            print(f"{category}: {total}₪")
        return categories

    def is_over_limit(self):
        if self.total_spent() > self.limitation:
            print("Budget exceeded! ❌")
            return True
        print("You are within the budget ❤")
        return False


# Ask user for budget limit
while True:
    try:
        limit = float(input("Enter your monthly budget limit:"))
        break
    except ValueError:
        print("Invalid number, please enter a valid amount.")

my_budget = Budget(limit)
my_budget.load_from_excel()
print("\n===== EXPENSES LOADED =====")
pprint(my_budget.expenses)

print("\n===== TOTAL SPENT =====")
print(f"{my_budget.total_spent()} ₪")

print("\n===== CATEGORY REPORT =====")
my_budget.category_report()

print("\n===== BUDGET STATUS =====")
my_budget.is_over_limit()


